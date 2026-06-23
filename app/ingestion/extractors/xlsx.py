import io

from openpyxl import load_workbook

from app.ingestion.extractors.base import BaseExtractor, ExtractedChunk

ROWS_PER_CHUNK = 50  # grouping per N baris


class XLSXExtractor(BaseExtractor):
    """
    Ekstrak konten XLSX per sheet, dikelompokkan setiap ROWS_PER_CHUNK baris.
    Menyimpan baris pertama (header) dan menempelkannya di setiap chunk
    agar konteks kolom (schema tabel) tidak hilang untuk RAG/LLM.
    """

    def extract(self, file_bytes: bytes, filename: str) -> list[ExtractedChunk]:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        chunks: list[ExtractedChunk] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_data: list[str] = []

            header_text = ""
            row_num = 0
            data_start_row = 0

            for row in ws.iter_rows(values_only=True):
                row_num += 1

                # Bersihkan nilai None dan string kosong
                cell_texts = [
                    str(c).strip()
                    for c in row
                    if c is not None and str(c).strip() != ""
                ]
                if not cell_texts:
                    continue  # Skip baris yang kosong sepenuhnya

                row_string = " | ".join(cell_texts)

                # Jadikan baris valid pertama sebagai Header
                if not header_text:
                    header_text = row_string
                    continue

                # Jika sudah ada header, catat baris ini sebagai awal data chunk
                if not rows_data:
                    data_start_row = row_num

                rows_data.append(row_string)

                # Flush setiap ROWS_PER_CHUNK baris data
                if len(rows_data) >= ROWS_PER_CHUNK:
                    # Tempelkan header di atas baris data
                    chunk_text = f"{header_text}\n" + "\n".join(rows_data)
                    chunks.append(
                        ExtractedChunk(
                            text=chunk_text.strip(),
                            sheet_name=sheet_name,
                            row_range=f"{data_start_row}-{row_num}",
                            metadata={"filename": filename},
                        )
                    )
                    rows_data = []

            # Flush sisa baris data yang kurang dari ROWS_PER_CHUNK
            if rows_data:
                chunk_text = f"{header_text}\n" + "\n".join(rows_data)
                chunks.append(
                    ExtractedChunk(
                        text=chunk_text.strip(),
                        sheet_name=sheet_name,
                        row_range=f"{data_start_row}-{row_num}",
                        metadata={"filename": filename},
                    )
                )
            # Edge case: Jika sheet hanya berisi 1 baris header saja tanpa data
            elif not chunks and header_text:
                chunks.append(
                    ExtractedChunk(
                        text=header_text,
                        sheet_name=sheet_name,
                        row_range="1-1",
                        metadata={"filename": filename},
                    )
                )

        wb.close()
        return chunks
